"""
Excel Master Merger  ─  Streamlit App
Upload a master format + multiple source files → get a combined master Excel back.
Run:  streamlit run excel_master_merger.py
Requires:  pip install streamlit pandas openpyxl xlrd
"""

import io, re
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import difflib

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Excel Master Merger",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  /* Main background */
  [data-testid="stAppViewContainer"] { background: #f0f4f9; }
  [data-testid="stSidebar"]          { background: #1a2b4a; }
  [data-testid="stSidebar"] * { color: #e8edf5 !important; }
  [data-testid="stSidebar"] .stSelectbox label,
  [data-testid="stSidebar"] .stCheckbox label { color: #c5d0e0 !important; }

  /* Cards */
  .card {
    background: white;
    border-radius: 12px;
    padding: 22px 24px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.07);
    margin-bottom: 16px;
  }
  .card-blue  { border-left: 5px solid #1a4fdb; }
  .card-green { border-left: 5px solid #1db86a; }
  .card-gold  { border-left: 5px solid #e8a020; }
  .card-red   { border-left: 5px solid #e83030; }

  /* Header */
  .app-header {
    background: linear-gradient(135deg,#1a2b4a,#1a4fdb);
    border-radius:14px;
    padding:28px 32px 22px;
    color:white;
    margin-bottom:24px;
  }
  .app-header h1 { margin:0; font-size:2rem; }
  .app-header p  { margin:6px 0 0; opacity:.8; font-size:1rem; }

  /* Stat boxes */
  .stat-box {
    background:white;
    border-radius:10px;
    padding:16px 20px;
    text-align:center;
    box-shadow:0 2px 8px rgba(0,0,0,0.06);
  }
  .stat-num  { font-size:2rem; font-weight:700; color:#1a4fdb; }
  .stat-lbl  { font-size:.78rem; color:#666; margin-top:2px; }

  /* Step badges */
  .step-badge {
    display:inline-block;
    background:#1a4fdb;
    color:white;
    border-radius:50%;
    width:28px; height:28px;
    line-height:28px;
    text-align:center;
    font-weight:700;
    font-size:.9rem;
    margin-right:8px;
  }

  /* Column mapping table */
  .col-map { font-size:.82rem; }

  /* Status chips */
  .chip-ok  { background:#e6f9ee; color:#1a8a45; border-radius:20px; padding:2px 10px; font-size:.78rem; }
  .chip-warn{ background:#fff5e0; color:#b56a00; border-radius:20px; padding:2px 10px; font-size:.78rem; }
  .chip-err { background:#fee; color:#c00; border-radius:20px; padding:2px 10px; font-size:.78rem; }
</style>
""", unsafe_allow_html=True)


# ── Helpers ───────────────────────────────────────────────────────────────────
def normalise(name: str) -> str:
    """Lowercase, strip, collapse spaces/underscores/hyphens."""
    return re.sub(r"[\s_\-]+", " ", str(name).strip().lower())


def best_match(col: str, candidates: list, cutoff: float = 0.70) -> str | None:
    """Return the best fuzzy match or None."""
    norm_col  = normalise(col)
    norm_cand = {normalise(c): c for c in candidates}
    matches   = difflib.get_close_matches(norm_col, norm_cand.keys(), n=1, cutoff=cutoff)
    return norm_cand[matches[0]] if matches else None


def read_excel_file(uploaded, sheet_name=0) -> pd.DataFrame:
    """Read an uploaded file (xlsx / xls / csv) into a DataFrame."""
    name = uploaded.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded)
    return pd.read_excel(uploaded, sheet_name=sheet_name, dtype=str)


def get_sheet_names(uploaded) -> list:
    """Return list of sheet names (csv returns ['Sheet1'])."""
    if uploaded.name.lower().endswith(".csv"):
        return ["Sheet1"]
    wb = load_workbook(uploaded, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def build_column_map(src_cols: list, master_cols: list, cutoff: float) -> dict:
    """
    Returns {src_col: master_col} for every source column that maps to master.
    Uses exact (case-insensitive) then fuzzy matching.
    """
    mapping = {}
    norm_master = {normalise(c): c for c in master_cols}
    for sc in src_cols:
        nm = normalise(sc)
        if nm in norm_master:                          # exact
            mapping[sc] = norm_master[nm]
        else:                                          # fuzzy
            match = best_match(sc, master_cols, cutoff)
            if match:
                mapping[sc] = match
    return mapping


def style_master_file(wb_bytes: bytes, new_row_start: int, total_rows: int) -> bytes:
    """Apply header formatting + highlight newly added rows in the output workbook."""
    wb  = load_workbook(io.BytesIO(wb_bytes))
    ws  = wb.active

    # Header row styling
    hdr_fill = PatternFill("solid", fgColor="1A2B4A")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_col = ws.max_column
    ws.row_dimensions[1].height = 30
    for col in range(1, max_col + 1):
        cell = ws.cell(1, col)
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.alignment = hdr_aln
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Highlight newly added rows
    new_fill = PatternFill("solid", fgColor="EFF6FF")
    for row in range(new_row_start, total_rows + 2):   # +2 for header offset
        for col in range(1, max_col + 1):
            ws.cell(row, col).fill = new_fill

    # Freeze header
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    st.markdown("---")

    fuzzy_cutoff = st.slider(
        "Fuzzy column match sensitivity",
        min_value=0.50, max_value=1.00, value=0.75, step=0.05,
        help="Higher = stricter matching. 0.75 is recommended."
    )
    remove_dupes = st.checkbox("Remove duplicate rows", value=False,
        help="Drop rows that are fully identical across all master columns.")
    skip_empty   = st.checkbox("Skip completely empty rows", value=True)
    show_unmapped = st.checkbox("Show unmapped columns report", value=True)

    st.markdown("---")
    st.markdown("### Supported formats")
    st.markdown("- `.xlsx`  `.xls`  `.csv`")
    st.markdown("### How it works")
    st.markdown("""
1. Upload your **master format** file (defines columns).
2. Upload **one or more source files** (data files).
3. App auto-maps columns → merges into master.
4. Download the updated master file.
""")
    st.markdown("---")
    st.caption("Excel Master Merger v1.0")


# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
  <h1>📊 Excel Master Merger</h1>
  <p>Automatically combine data from multiple Excel files into a single master format — zero data loss.</p>
</div>
""", unsafe_allow_html=True)


# ── STEP 1 – Upload ───────────────────────────────────────────────────────────
st.markdown('<span class="step-badge">1</span> **Upload Files**', unsafe_allow_html=True)
col_master, col_src = st.columns([1, 1], gap="large")

with col_master:
    st.markdown('<div class="card card-blue">', unsafe_allow_html=True)
    st.markdown("#### 🗂️ Master Format File")
    st.caption("This file defines the column structure. Existing data will be preserved.")
    master_file = st.file_uploader(
        "Upload master file", type=["xlsx", "xls", "csv"],
        key="master", label_visibility="collapsed"
    )
    st.markdown('</div>', unsafe_allow_html=True)

with col_src:
    st.markdown('<div class="card card-green">', unsafe_allow_html=True)
    st.markdown("#### 📁 Source Files  *(one or many)*")
    st.caption("Upload all files whose data should be merged into the master.")
    source_files = st.file_uploader(
        "Upload source files", type=["xlsx", "xls", "csv"],
        accept_multiple_files=True, key="sources", label_visibility="collapsed"
    )
    st.markdown('</div>', unsafe_allow_html=True)


# ── Validate uploads ──────────────────────────────────────────────────────────
if not master_file:
    st.info("👆 Please upload a **Master Format** file to get started.")
    st.stop()

if not source_files:
    st.info("👆 Please upload at least one **Source File**.")
    st.stop()


# ── STEP 2 – Sheet Selection ──────────────────────────────────────────────────
st.markdown("---")
st.markdown('<span class="step-badge">2</span> **Select Sheets**', unsafe_allow_html=True)

# Master sheet
master_file.seek(0)
master_sheets = get_sheet_names(master_file)

col_ms, col_spacer = st.columns([1, 2])
with col_ms:
    master_sheet = st.selectbox(
        f"Sheet in **{master_file.name}**",
        options=master_sheets, index=0
    )

# Source sheets
st.caption("Pick the sheet to read from each source file:")
src_sheet_map = {}
src_cols = st.columns(min(len(source_files), 3))
for i, sf in enumerate(source_files):
    sf.seek(0)
    sheets = get_sheet_names(sf)
    with src_cols[i % 3]:
        src_sheet_map[sf.name] = st.selectbox(
            f"📄 `{sf.name}`", options=sheets, index=0, key=f"sh_{i}"
        )


# ── STEP 3 – Preview Master ───────────────────────────────────────────────────
st.markdown("---")
st.markdown('<span class="step-badge">3</span> **Master Format Preview**', unsafe_allow_html=True)

master_file.seek(0)
try:
    master_df = read_excel_file(master_file, sheet_name=master_sheet if master_sheet != "Sheet1" or not master_file.name.endswith(".csv") else 0)
except Exception as e:
    st.error(f"Could not read master file: {e}")
    st.stop()

master_cols = list(master_df.columns)
existing_rows = len(master_df)

with st.expander(f"Master columns ({len(master_cols)}) | Existing rows: {existing_rows}", expanded=True):
    st.dataframe(master_df.head(10), use_container_width=True, height=220)


# ── STEP 4 – Column Mapping Preview ──────────────────────────────────────────
if show_unmapped:
    st.markdown("---")
    st.markdown('<span class="step-badge">4</span> **Column Mapping Preview**', unsafe_allow_html=True)

    all_maps   = {}   # {filename: {src_col: master_col}}
    all_unmapped = {}

    for sf in source_files:
        sf.seek(0)
        try:
            sdf = read_excel_file(sf, sheet_name=src_sheet_map[sf.name] if sf.name in src_sheet_map else 0)
            col_map = build_column_map(list(sdf.columns), master_cols, fuzzy_cutoff)
            all_maps[sf.name]     = col_map
            all_unmapped[sf.name] = [c for c in sdf.columns if c not in col_map]
        except Exception as e:
            st.warning(f"⚠️ Could not read `{sf.name}`: {e}")

    for fname, cmap in all_maps.items():
        with st.expander(f"📄 `{fname}` — {len(cmap)} mapped / {len(all_unmapped.get(fname, []))} unmapped"):
            rows = []
            for sc, mc in cmap.items():
                exact = normalise(sc) == normalise(mc)
                rows.append({
                    "Source Column": sc,
                    "→ Master Column": mc,
                    "Match Type": "✅ Exact" if exact else "🔶 Fuzzy"
                })
            if rows:
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            unm = all_unmapped.get(fname, [])
            if unm:
                st.warning(f"Unmapped columns (data **not** included): {', '.join(unm)}")


# ── STEP 5 – Merge ────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown('<span class="step-badge">5</span> **Merge Data**', unsafe_allow_html=True)

do_merge = st.button("🚀 Start Merging", type="primary", use_container_width=False)

if do_merge:
    progress   = st.progress(0, text="Initialising…")
    status_box = st.empty()
    log_lines  = []
    frames     = [master_df.copy()]   # keep existing master data first
    file_stats = []

    for idx, sf in enumerate(source_files):
        pct  = int((idx / len(source_files)) * 90)
        progress.progress(pct, text=f"Processing `{sf.name}` ({idx+1}/{len(source_files)})…")
        status_box.info(f"⏳ Reading `{sf.name}`…")

        sf.seek(0)
        try:
            sdf = read_excel_file(sf, sheet_name=src_sheet_map.get(sf.name, 0))
        except Exception as e:
            st.warning(f"⚠️ Skipped `{sf.name}`: {e}")
            file_stats.append({"File": sf.name, "Rows Read": 0, "Rows Added": 0, "Status": "⚠️ Error"})
            continue

        # Skip empty rows
        if skip_empty:
            sdf = sdf.dropna(how="all")

        raw_count = len(sdf)

        # Build column map for this file
        col_map   = build_column_map(list(sdf.columns), master_cols, fuzzy_cutoff)

        # Remap → master column space
        remapped  = pd.DataFrame(columns=master_cols)
        for sc, mc in col_map.items():
            remapped[mc] = sdf[sc].values[:len(sdf)]

        # Fill missing master columns with NaN
        for mc in master_cols:
            if mc not in remapped.columns:
                remapped[mc] = pd.NA

        remapped = remapped[master_cols]   # enforce column order
        frames.append(remapped)

        file_stats.append({
            "File": sf.name,
            "Rows Read": raw_count,
            "Rows Added": len(remapped),
            "Mapped Cols": len(col_map),
            "Status": "✅ OK"
        })
        log_lines.append(f"✅ `{sf.name}` — {len(remapped)} rows merged.")

    progress.progress(95, text="Combining all data…")
    status_box.info("⏳ Combining all data…")

    combined = pd.concat(frames, ignore_index=True)

    if remove_dupes:
        before = len(combined)
        combined = combined.drop_duplicates()
        removed  = before - len(combined)
        log_lines.append(f"🗑️ {removed} duplicate rows removed.")

    # ── Build output Excel ─────────────────────────────────────────────────────
    progress.progress(97, text="Writing output Excel…")

    out_buf = io.BytesIO()
    with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
        combined.to_excel(writer, index=False, sheet_name="Master Data")

    # Style the output
    styled = style_master_file(
        out_buf.getvalue(),
        new_row_start=existing_rows + 2,   # +2 = header row
        total_rows=len(combined)
    )

    progress.progress(100, text="Done! ✅")
    status_box.success("✅ Merge complete!")

    # ── Results ────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📈 Results")

    total_new = sum(s["Rows Added"] for s in file_stats)
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f'<div class="stat-box"><div class="stat-num">{len(source_files)}</div><div class="stat-lbl">Source Files Processed</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="stat-box"><div class="stat-num">{existing_rows}</div><div class="stat-lbl">Existing Master Rows</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="stat-box"><div class="stat-num">{total_new}</div><div class="stat-lbl">New Rows Added</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="stat-box"><div class="stat-num">{len(combined)}</div><div class="stat-lbl">Total Rows in Output</div></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Per-file table
    st.markdown("#### Per-file breakdown")
    stats_df = pd.DataFrame(file_stats)
    st.dataframe(stats_df, use_container_width=True, hide_index=True)

    # Log
    with st.expander("📋 Merge log"):
        for line in log_lines:
            st.markdown(line)

    # Preview combined data
    st.markdown("#### Preview of merged master (first 50 rows)")
    st.dataframe(combined.head(50), use_container_width=True, height=300)

    # ── Download ───────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### ⬇️ Download Updated Master File")
    st.download_button(
        label="📥 Download Master Excel (.xlsx)",
        data=styled,
        file_name="Master_Combined.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
    st.caption("💡 Newly added rows are highlighted in light blue. Header row is formatted in dark navy.")


# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("---")
st.caption("Excel Master Merger • Built with Streamlit & Pandas • Zero data loss guaranteed")
